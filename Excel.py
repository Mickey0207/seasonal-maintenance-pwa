import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 新建活頁簿與工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "1008-範本"

# 設定主要標題列
header = [
    "決裁編號", "", "", "", "", "", "工程名稱", "", "MITSUI OUTLET PARK 台南", "", "", "", "", "", "", "", "", "", "", "驗收照片", "", "", "", "", "", "", "", "", "日期", "",
    "112.03.09~\n112.03.13"
]
ws.append(header)

# 合併儲存格（依原檔合併規格，可依實際需求細調）
ws.merge_cells('A1:F1')
ws.merge_cells('G1:H1')
ws.merge_cells('I1:T1')
ws.merge_cells('U1:AD1')
ws.merge_cells('AE1:AF1')
ws.merge_cells('AG1:AH1')

# 設定標題字型與填色
header_font = Font(size=12, bold=True, name='Arial')
header_fill = PatternFill("solid", fgColor="FBD4B4")
for col in range(1, 33):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 假設第二列起為區塊標題
ws.append(['', '', '№１', '', '', '', '', '', '№２', '', '', '', '', '', '№３', '', '', '', '', '', '№４'])

# 設定編號列格式
block_font = Font(size=11, bold=True)
for i in [3, 9, 15, 21]:
    ws.cell(row=2, column=i).font = block_font
    ws.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center')

# 增加一組「位置」「内容」「備註」區塊
for offset in [3, 9, 15, 21]:
    ws.cell(row=3, column=offset, value="位置：").alignment = Alignment(horizontal='left')
    ws.cell(row=4, column=offset, value="内容：").alignment = Alignment(horizontal='left')
    ws.cell(row=5, column=offset, value="備註：").alignment = Alignment(horizontal='left')

# 設定邊框
thin = Side(border_style='thin', color='000000')
for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=32):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 可依範本結構迴圈繼續寫入表格，其餘略

# 儲存檔案
wb.save('範本_手動打造.xlsx')
